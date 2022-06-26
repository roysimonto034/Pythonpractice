import os
import sys
import openpyxl
from openpyxl.styles import Font, Color
import pytest
import logging
import subprocess
import datetime
from collections import deque
from openpyxl.styles import Border, PatternFill, Side, Alignment, Protection, Font, colors
"""pytest -vv mfpdevice/capability/>test_result.txt"""

logging.basicConfig(filename=os.path.join(os.path.dirname(os.path.abspath(__file__)),"test_utr_auto.log"), 
                    format='%(asctime)s %(message)s', 
                    filemode='w') 
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)


green_background = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
red_background = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')
fnt=Font(size=14,bold=True)
alg=Alignment(horizontal='center')
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


def auto_script():
    # t1=time.clock()
    t1=datetime.datetime.now()
    module = sys.argv[1]
    #import pdb;pdb.set_trace()
    udf,mdf,gl,fundct,cmnt,quess,fildct,expct=[],[],[],{},{},{},{},[]
    # subprocess.call(["export"," ","PATH=$PATH:/home/SYSROM_SRC/build/thirdparty/bin/"])
    #subprocess.call(["pytest"])
    # ps=subprocess.call(["pytest","-vv",module])
    # test_status={}
    path=os.path.dirname(os.path.abspath(__file__))
    logger.info("path for file execution: "+str(path))
    with open(os.path.join(path,'test_result.txt'),'r')as f:
        for ip in f.readlines():
            # if ip.startswith(module.split('/')[0]):
                # if not 'Error' in ip:
            if '::' in ip:
                if 'test_' in ip:
                    if (']') in ip:
                        if ip.split('::')[1].split(']')[1].startswith(' PASS'):
                           # import pdb;pdb.set_trace()
                           udf.append((ip.split('::')[0].split('/')[-1]+'%'+ip.split('::')[1].split(']')[0].replace('[','_'),'P'))
                           # fildct[ip.split('::')[0]]=(ip.split('::')[1].split(']')[0].replace('[','_'),'P')
                        elif ip.split('::')[1].split(']')[1].startswith(' FAIL'):
                             udf.append((ip.split('::')[0].split('/')[-1]+'%'+ip.split('::')[1].split(']')[0].replace('[','_'),'F'))
                             # fildct[ip.split('::')[0]]=(ip.split('::')[1].split(']')[0].replace('[','_'),'F')
                             #mdf.append((ip.split('::')[0],ip.split('::')[1]))
                        elif ip.split('::')[1].split(']')[1].startswith(' XPASS'):
                             udf.append((ip.split('::')[0].split('/')[-1]+'%'+ip.split('::')[1].split(']')[0].replace('[','_'),'F'))
                             # fildct[ip.split('::')[0]]=(ip.split('::')[1].split(']')[0].replace('[','_'),'F')
                        else:
                           udf.append((ip.split('::')[0].split('/')[-1]+'%'+ip.split('::')[1].split(']')[0].replace('[','_'),'P'))
                           # fildct[ip.split('::')[0]]=(ip.split('::')[1].split(']')[0].replace('[','_'),'P')
                    else:
                        #import pdb;pdb.set_trace()
                        if ip.split('::')[1].split(' ')[1].startswith('PASS'):
                           udf.append((ip.split('::')[0].split('/')[-1]+'%'+ip.split('::')[1].split(' ')[0].replace('[','_'),'P'))
                           # fildct[ip.split('::')[0]]=(ip.split('::')[1].split(' ')[0].replace('[','_'),'P')
                        elif ip.split('::')[1].split(' ')[1]. startswith('FAIL'):
                             udf.append((ip.split('::')[0].split('/')[-1]+'%'+ip.split('::')[1].split(' ')[0].replace('[','_'),'F'))
                             # fildct[ip.split('::')[0]]=(ip.split('::')[1].split(' ')[0].replace('[','_'),'F')
							#mdf.append((ip.split('::')[0],ip.split('::')[1]))
                        elif ip.split('::')[1].split(' ')[1]. startswith('XPASS'):
                             udf.append((ip.split('::')[0].split('/')[-1]+'%'+ip.split('::')[1].split(' ')[0].replace('[','_'),'F'))
                             # fildct[ip.split('::')[0]]=(ip.split('::')[1].split(' ')[0].replace('[','_'),'F')
                        else:
                            udf.append((ip.split('::')[0].split('/')[-1]+'%'+ip.split('::')[1].split(' ')[0].replace('[','_'),'P'))
                            # fildct[ip.split('::')[0]]=(ip.split('::')[1].split(' ')[0].replace('[','_'),'P')

    file=os.listdir(os.path.join(path,module))
    for fl in file:
        if fl[:4]=='test':
            gl.append(fl)
    print(fl)
    for fp in gl:
        ml,test_function,comments,obj,stp,temp_det,my_dc,fin_lst=[],[],[],[],'',[],{},[]
        expected = ['1.Request objects are mocked.','2.Test returns the actual expected result']
        dcp={'assert_called_with':'assert is called with ()','raises': 'is expected to raise exception','raise_eapi_exception':'is expected to raise eapi exception'}
        with open(os.path.join(path,module,fp),'r') as f:
            ml=f.readlines()
            tx=[ml.index(x) for x in ml if x.startswith('def')]
            gc,gt,fg,lg,gl,gx,cmnt,exp=deque([]),[],[],[],[],[],{},[]
            gc.append(tx[0])
            dp=deque(tx)
            dp.popleft()
            for ip in range(len(tx)-1):
                gc.append(dp.popleft())
                # import pdb;pdb.set_trace()
                if len(gc)==2:
                    for i in range(gc[0],gc[1]):
                        # import pdb;pdb.set_trace()
                        gt.append(ml[i])
                    fg.append(gt)
                    gc.popleft()
                    gt=[]
            for ip in range(gc[0],len(ml)-1):
                gx.append(ml[ip])
            fg.append(gx)
            #fg.append(gx)
            for pl in fg:
                if pl[0].startswith('def test'):
                    gl.append(pl)
            for tl in gl:
                for val in tl:
                    # import pdb;pdb.set_trace()
                    if val.startswith('#') or val.startswith(' \n') or val.startswith('    #') or val.startswith('@') or val.startswith('from') or val.startswith('import') or val.startswith('\n')or val.startswith('([') or val.startswith('])\n') or val.startswith('def mock'):
                        continue
                    else:
                        # import pdb;pdb.set_trace()
                        if val.startswith('def test'):
                            if test_function:
                                test_function.pop()
                            test_function.append(val.split('(')[0].split(' ')[1])
                            cmnt={}
                        elif val.startswith('    """Test case Objective:'):
                            cmnt['Test case Objective']=val.split(': ')[1].strip('\n')
                        elif val.startswith('    Precondition:'):
                            cmnt['Precondition']=val.split(': ')[1].strip('\n')
                        elif val.startswith('    Expected Result:'):
                            # import pdb;pdb.set_trace()
                            value=val.split(': ')[1].strip('"""\n').strip('[]')
                            if value:
                                    cmnt['Expected Result']=[]
                                    cmnt['Expected Result'].append(value)
                                    name = fp + '%' + test_function[0]
                                    fundct[name] = cmnt
                            else:
                                cmnt['Expected Result']=[]
                        elif 'return_value' in val and ' ' not in cmnt['Expected Result']:
                                # import pdb;pdb.set_trace()
                                spl = val.split('=')[0]
                                if '(' in spl:
                                    meth = spl.split('.')[-3]
                                else:
                                    meth = spl.split('.')[-2]
                                cmnt['Expected Result'].append(meth+" is mocked")
                                name = fp + '%' + test_function[0]
                                fundct[name] = cmnt
                        elif 'assert_called_with' in val or 'assert_called_once_with' in val and ' ' not in cmnt['Expected Result']:
                                # import pdb;pdb.set_trace()
                                met = val.split('(')[1].strip('\n').strip(')')
                                if met:
                                    try:
                                        # import pdb;pdb.set_trace()
                                        cmnt['Expected Result'].append("assert raised with "+met)
                                        name = fp + '%' + test_function[0]
                                        fundct[name] = cmnt
                                    except AttributeError as e:
                                        # import pdb;pdb.set_trace()
                                        print(cmnt)
                                else:
                                    cmnt['Expected Result'].append("assert raised with None")
    for ig in tuple(fundct.keys()):
        #import pdb;pdb.set_trace()
        for uf in udf:
            if uf[0].startswith(ig):
                # import pdb;pdb.set_trace()
                quess[uf]=fundct[ig]
            else:
                continue
    wb = openpyxl.Workbook()
    name = wb.get_sheet_names()
    splt=module.split("/")
    if len(splt)>2:
        if splt[-1]=='':
            titl = splt[-2]
        else:
            titl = splt[-1]
    else:
        if splt[-1]:
            titl=splt[-1]
        else:
            titl=splt[-2]
    wb.create_sheet(index = 1,title = titl)
    header=['Sl.No','File Name','Test ID','Test case Objective','Precondition',\
            'Test Data','Test Procedure','Expected Result','Actual Result','Status']
    sheet = wb.get_sheet_by_name(titl)
    for ln in range(1,len(header)+1):
        for cn in range(1,len(udf)+2):
            sheet.cell(row =cn, column = ln).border=thin_border
    for cn in range(1,len(header)+1):
        sheet.cell(row = 1, column = cn).font=fnt
        sheet.cell(row = 1, column = cn).alignment=alg
        sheet.cell(row = 1, column = cn).value=header[cn-1]
    for cp in range(1,len(udf)+1):
        sheet.cell(row = cp+1, column = 1).alignment=alg
        sheet.cell(row = cp+1, column = 1).value=cp
    for gn in range(1,len(quess.keys())+1):
        sheet.cell(row =gn+1, column = 2).value= tuple(quess.keys())[gn-1][0].split('%')[0]
    for gp in range(1,len(quess.keys())+1):
        sheet.cell(row =gp+1, column = 3).value= tuple(quess.keys())[gp-1][0].split('%')[1]
    for gr in range(1,len(quess.keys())+1):
        sheet.cell(row =gr+1, column = 4).value= quess[tuple(quess.keys())[gr-1]]['Test case Objective']
    for gs in range(1,len(quess.keys())+1):
        sheet.cell(row =gs+1, column = 5).value= quess[tuple(quess.keys())[gs-1]]['Precondition'].strip('(')
    for gt in range(1,len(quess.keys())+1):
        # sheet.cell(row =gt+1, column = 8).alignment.wrap_text = Alignment(wrap_text=True)
        sheet.cell(row =gt+1, column = 8).value= str(quess[tuple(quess.keys())[gt-1]]['Expected Result']).strip('[]').strip("'").strip('[]').strip(',').strip("',")
    for ig in range(1,len(quess.keys())+1):
        sheet.cell(row =ig+1, column = 10).alignment=alg
        sheet.cell(row =ig+1, column = 10).font=Font(size=12,name='Times New Roman')
        sheet.cell(row =ig+1, column = 10).value=tuple(quess.keys())[ig-1][1]
    for i in range(1,len(udf)+1):
        if sheet.cell(row=i+1,column=10).value=='P':
            sheet.cell(row=i+1,column=10).fill=green_background
			# sheet.cell(row=i+1,column=10).style.alignment.wrap_text = True
            sheet.cell(row=i+1,column=9).value=sheet.cell(row=i+1,column=8).value
        elif sheet.cell(row=i+1,column=10).value=='F':
            sheet.cell(row=i+1,column=10).fill=red_background
    wb.save(os.path.join(path,"6feb_Unit_Test_Report.xlsx"))
    t2=datetime.datetime.now()
    # print("time taken by code: ",t2-t1)
    diff=t2-t1
    # print(time.strftime("%H:%M:%S", time.gmtime(t2-t1)))
    print(diff.seconds,diff.microseconds)
    # return udf 4,5,8



if __name__=="__main__":
	auto_script()
